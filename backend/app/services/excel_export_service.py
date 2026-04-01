"""엑셀 다운로드 서비스 (REQ-005).

1단계: 원본 요구사항만 4컬럼으로 출력.
2단계: 원본 + 상세를 7컬럼으로 출력하며 원본 열(A-D)을 상세 행 수만큼 세로 병합한다.
"""

import io
from collections import defaultdict
from typing import Literal

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import app.state as state

# 헤더 행 스타일 — 진한 파랑 배경, 흰 글자
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# 원본 열(A-D) 배경색 — 연한 파랑
ORIG_FILL = PatternFill("solid", fgColor="D9E1F2")

# 수정된 상세 행 배경색 — 연한 노랑 (is_modified=True인 G열에 적용)
MOD_FILL = PatternFill("solid", fgColor="FFF2CC")


class ExcelExportService:
    """SessionStore 데이터를 openpyxl로 변환하여 xlsx 바이너리를 반환하는 서비스."""

    def export(self, stage: Literal[1, 2]) -> bytes:
        """요청 stage에 따라 1단계 또는 2단계 엑셀을 생성하고 바이너리를 반환한다.

        Args:
            stage: 1 = 원본 요구사항만, 2 = 원본 + 상세 통합 레이아웃

        Returns:
            xlsx 파일 바이너리

        Raises:
            HTTPException 422: stage=2인데 상세요구사항이 없는 경우
        """
        originals = state.get_original()

        if stage == 2:
            details = state.get_detail()
            if not details:
                raise HTTPException(
                    422,
                    detail={
                        "code": "DETAIL_NOT_GENERATED",
                        "message": "상세요구사항을 먼저 생성해주세요.",
                    },
                )

        wb = Workbook()
        ws = wb.active

        if stage == 1:
            _write_stage1(ws, originals)
        else:
            _write_stage2(ws, originals, state.get_detail())

        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def _write_stage1(ws, originals: list) -> None:
    """1단계 시트를 작성한다 — 원본 요구사항 4컬럼."""
    ws.title = "원본요구사항"
    headers = ["요구사항 ID", "분류", "요구사항 명칭", "요구사항 내용"]
    col_widths = [15, 12, 30, 60]
    _write_header(ws, headers, col_widths)

    for i, req in enumerate(originals, start=2):
        ws.cell(i, 1, req.id)
        ws.cell(i, 2, req.category)
        ws.cell(i, 3, req.name)
        ws.cell(i, 4, req.content).alignment = Alignment(wrap_text=True)
        for c in range(1, 5):
            ws.cell(i, c).fill = ORIG_FILL


def _write_stage2(ws, originals: list, details: list) -> None:
    """2단계 시트를 작성한다 — 원본(A-D) + 상세(E-G) 병합 레이아웃.

    원본 1건당 상세 N건을 우측 열에 나열하고,
    원본 열(A-D)은 상세 행 수만큼 세로 병합하여 시각적으로 묶는다.
    """
    ws.title = "상세요구사항"
    headers = [
        "원본 요구사항 ID", "분류", "원본 명칭", "원본 내용",
        "상세 요구사항 ID", "상세 명칭", "상세 내용",
    ]
    col_widths = [15, 12, 30, 50, 15, 30, 60]
    _write_header(ws, headers, col_widths)

    # parent_id 기준으로 상세 그룹핑
    detail_map: dict[str, list] = defaultdict(list)
    for d in details:
        detail_map[d.parent_id].append(d)

    row = 2
    for orig in originals:
        dets = detail_map.get(orig.id, [])
        count = max(len(dets), 1)
        start_row = row

        # 원본 데이터는 병합될 첫 행에만 기록한다
        ws.cell(row, 1, orig.id)
        ws.cell(row, 2, orig.category)
        ws.cell(row, 3, orig.name)
        ws.cell(row, 4, orig.content).alignment = Alignment(wrap_text=True)

        # 상세 데이터를 행별로 기록한다
        for det in dets:
            ws.cell(row, 5, det.id)
            ws.cell(row, 6, det.name)
            cell = ws.cell(row, 7, det.content)
            cell.alignment = Alignment(wrap_text=True)
            # 수정된 상세 내용은 노란 배경으로 구분한다
            if det.is_modified:
                cell.fill = MOD_FILL
            row += 1

        if not dets:
            row += 1

        # 상세가 2개 이상인 경우 원본 열(A-D)을 세로 병합한다
        if count > 1:
            for col in range(1, 5):
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=start_row + count - 1,
                    end_column=col,
                )
                ws.cell(start_row, col).alignment = Alignment(
                    vertical="top", wrap_text=(col == 4)
                )

        # 원본 열 배경색 적용
        for r in range(start_row, start_row + count):
            for c in range(1, 5):
                ws.cell(r, c).fill = ORIG_FILL


def _write_header(ws, headers: list[str], widths: list[int]) -> None:
    """헤더 행을 작성하고 열 너비를 설정한다.

    Args:
        ws: 대상 워크시트
        headers: 헤더 텍스트 목록
        widths: 열 너비 목록 (문자 단위)
    """
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = w
