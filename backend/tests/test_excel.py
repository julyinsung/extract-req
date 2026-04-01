"""엑셀 다운로드 서비스 단위 테스트.

UT-005-01: export(stage=1) → 4컬럼 xlsx, 원본 행 수 일치
UT-005-02: export(stage=2) → 7컬럼 xlsx, 원본 셀 병합 확인
UT-005-03: GET /api/v1/download → Content-Type xlsx 헤더 확인
UT-005-04: is_modified=True 행 → 2단계 엑셀에 수정값과 MOD_FILL 적용
UT-005-05: stage=2 + detailReqs 없음 → HTTPException 422
"""

import io
import json
import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import app.state as state
from app.main import app
from app.models.requirement import DetailRequirement, OriginalRequirement
from app.services.excel_export_service import ExcelExportService, MOD_FILL


# ---------------------------------------------------------------------------
# 픽스처
# ---------------------------------------------------------------------------


def _make_originals() -> list[OriginalRequirement]:
    """원본 요구사항 픽스처 — 2건."""
    return [
        OriginalRequirement(
            id="SFR-001",
            category="기능요구사항",
            name="파일 업로드",
            content="HWP 파일을 업로드할 수 있다.",
            order_index=0,
        ),
        OriginalRequirement(
            id="SFR-002",
            category="유지관리",
            name="결과 다운로드",
            content="엑셀로 다운로드할 수 있다.",
            order_index=1,
        ),
    ]


def _make_details() -> list[DetailRequirement]:
    """상세요구사항 픽스처 — SFR-001: 3건, SFR-002: 2건."""
    return [
        DetailRequirement(
            id="SFR-001-01",
            parent_id="SFR-001",
            category="기능요구사항",
            name="드래그앤드롭 업로드",
            content="드래그앤드롭으로 업로드한다.",
            order_index=0,
        ),
        DetailRequirement(
            id="SFR-001-02",
            parent_id="SFR-001",
            category="기능요구사항",
            name="파일 유효성 검사",
            content="HWP 형식만 허용한다.",
            order_index=1,
        ),
        DetailRequirement(
            id="SFR-001-03",
            parent_id="SFR-001",
            category="기능요구사항",
            name="업로드 진행 표시",
            content="진행률을 표시한다.",
            order_index=2,
        ),
        DetailRequirement(
            id="SFR-002-01",
            parent_id="SFR-002",
            category="유지관리",
            name="1단계 다운로드",
            content="원본 요구사항만 내보낸다.",
            order_index=0,
        ),
        DetailRequirement(
            id="SFR-002-02",
            parent_id="SFR-002",
            category="유지관리",
            name="2단계 다운로드",
            content="원본 + 상세 통합 내보내기.",
            order_index=1,
        ),
    ]


def _load_ws_from_bytes(data: bytes):
    """bytes에서 openpyxl 워크시트를 로드한다."""
    wb = load_workbook(io.BytesIO(data))
    return wb.active


# ---------------------------------------------------------------------------
# UT-005-01: export(stage=1) → 4컬럼 xlsx, 원본 행 수 일치
# ---------------------------------------------------------------------------


class TestExportStage1:
    """UT-005-01: 1단계 내보내기가 4컬럼으로 원본 데이터를 정확히 출력해야 한다."""

    def setup_method(self):
        state.reset_session()
        state.set_original(_make_originals())

    def test_stage1_has_four_columns(self):
        """1단계 엑셀은 4개 컬럼(A-D)만 포함해야 한다."""
        service = ExcelExportService()
        data = service.export(1)
        ws = _load_ws_from_bytes(data)

        # 헤더 행에서 실제 사용된 컬럼 수 확인
        header_cells = [ws.cell(1, c).value for c in range(1, 6)]
        filled = [h for h in header_cells if h is not None]
        assert len(filled) == 4, f"4개 헤더여야 하지만 {len(filled)}개다"

    def test_stage1_row_count_matches_originals(self):
        """데이터 행 수가 원본 요구사항 수와 일치해야 한다."""
        originals = _make_originals()
        service = ExcelExportService()
        data = service.export(1)
        ws = _load_ws_from_bytes(data)

        # 헤더(1행) 제외 데이터 행 수
        data_rows = sum(
            1 for row in ws.iter_rows(min_row=2) if any(c.value for c in row)
        )
        assert data_rows == len(originals), (
            f"데이터 행이 {len(originals)}개여야 하지만 {data_rows}개다"
        )

    def test_stage1_data_values_correct(self):
        """1단계 엑셀의 데이터 값이 원본 요구사항과 일치해야 한다."""
        service = ExcelExportService()
        data = service.export(1)
        ws = _load_ws_from_bytes(data)

        assert ws.cell(2, 1).value == "SFR-001"
        assert ws.cell(2, 2).value == "기능요구사항"
        assert ws.cell(2, 3).value == "파일 업로드"
        assert ws.cell(3, 1).value == "SFR-002"

    def test_stage1_sheet_name(self):
        """1단계 시트명이 '원본요구사항'이어야 한다."""
        service = ExcelExportService()
        data = service.export(1)
        wb = load_workbook(io.BytesIO(data))
        assert wb.active.title == "원본요구사항"


# ---------------------------------------------------------------------------
# UT-005-02: export(stage=2) → 7컬럼 xlsx, 원본 셀 병합 확인
# ---------------------------------------------------------------------------


class TestExportStage2:
    """UT-005-02: 2단계 내보내기가 7컬럼으로 원본 셀 병합 레이아웃을 올바르게 출력해야 한다."""

    def setup_method(self):
        state.reset_session()
        state.set_original(_make_originals())
        state.set_detail(_make_details())

    def test_stage2_has_seven_columns(self):
        """2단계 엑셀은 7개 컬럼(A-G)을 포함해야 한다."""
        service = ExcelExportService()
        data = service.export(2)
        ws = _load_ws_from_bytes(data)

        header_cells = [ws.cell(1, c).value for c in range(1, 9)]
        filled = [h for h in header_cells if h is not None]
        assert len(filled) == 7, f"7개 헤더여야 하지만 {len(filled)}개다"

    def test_stage2_merged_cells_exist(self):
        """원본 열(A-D)이 상세 행 수만큼 세로 병합되어야 한다."""
        service = ExcelExportService()
        data = service.export(2)
        ws = _load_ws_from_bytes(data)

        merged = ws.merged_cells.ranges
        assert len(merged) > 0, "병합 셀이 1개 이상 존재해야 한다"

    def test_stage2_sfr001_merged_three_rows(self):
        """SFR-001은 상세 3건이므로 A열이 3행 병합되어야 한다 (행 2~4)."""
        service = ExcelExportService()
        data = service.export(2)
        ws = _load_ws_from_bytes(data)

        # SFR-001의 A열 병합 범위 확인
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        assert any("A2:A4" in r for r in merged_ranges), (
            f"A2:A4 병합이 있어야 하지만 실제 병합: {merged_ranges}"
        )

    def test_stage2_sheet_name(self):
        """2단계 시트명이 '상세요구사항'이어야 한다."""
        service = ExcelExportService()
        data = service.export(2)
        wb = load_workbook(io.BytesIO(data))
        assert wb.active.title == "상세요구사항"

    def test_stage2_detail_order_correct(self):
        """상세 요구사항이 parent_id 순서대로 배치되어야 한다."""
        service = ExcelExportService()
        data = service.export(2)
        ws = _load_ws_from_bytes(data)

        # 행 2의 E열이 SFR-001의 첫 번째 상세여야 한다
        assert ws.cell(2, 5).value == "SFR-001-01"
        # 행 5의 E열이 SFR-002의 첫 번째 상세여야 한다 (SFR-001 3건 후)
        assert ws.cell(5, 5).value == "SFR-002-01"


# ---------------------------------------------------------------------------
# UT-005-03: GET /api/v1/download → Content-Type xlsx 헤더 확인
# ---------------------------------------------------------------------------


class TestDownloadEndpoint:
    """UT-005-03: GET /api/v1/download가 올바른 Content-Type을 반환해야 한다."""

    def setup_method(self):
        state.reset_session()
        state.set_original(_make_originals())

    def test_download_stage1_content_type(self):
        """GET /api/v1/download?stage=1 응답의 Content-Type이 xlsx여야 한다."""
        client = TestClient(app)
        response = client.get("/api/v1/download", params={"session_id": "test", "stage": 1})

        assert response.status_code == 200
        assert (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            in response.headers["content-type"]
        )

    def test_download_stage1_content_disposition(self):
        """응답 헤더에 attachment 다운로드 설정이 포함되어야 한다."""
        client = TestClient(app)
        response = client.get("/api/v1/download", params={"session_id": "test", "stage": 1})

        assert "attachment" in response.headers.get("content-disposition", "")
        assert "requirements_original_" in response.headers.get("content-disposition", "")

    def test_invalid_stage_returns_422(self):
        """stage=3 같은 잘못된 값은 422를 반환해야 한다."""
        client = TestClient(app)
        response = client.get("/api/v1/download", params={"session_id": "test", "stage": 3})
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# UT-005-04: is_modified=True 행 → 2단계 엑셀에 수정값 포함 및 MOD_FILL 적용
# ---------------------------------------------------------------------------


class TestModifiedFill:
    """UT-005-04: is_modified=True 행의 G열에 MOD_FILL이 적용되어야 한다."""

    def setup_method(self):
        state.reset_session()
        state.set_original(_make_originals()[:1])  # SFR-001만 사용

    def test_modified_cell_has_mod_fill(self):
        """is_modified=True인 상세요구사항의 G열에 노란 배경이 적용되어야 한다."""
        details = [
            DetailRequirement(
                id="SFR-001-01",
                parent_id="SFR-001",
                category="기능요구사항",
                name="수정된 명칭",
                content="수정된 내용",
                order_index=0,
                is_modified=True,
            )
        ]
        state.set_detail(details)

        service = ExcelExportService()
        data = service.export(2)
        ws = _load_ws_from_bytes(data)

        # 데이터 2행의 G열(상세 내용) 배경색 확인
        cell = ws.cell(2, 7)
        assert cell.fill.fgColor.rgb == MOD_FILL.fgColor.rgb, (
            f"수정된 셀의 배경색이 MOD_FILL({MOD_FILL.fgColor.rgb})이어야 하지만 "
            f"{cell.fill.fgColor.rgb}이다"
        )

    def test_unmodified_cell_no_mod_fill(self):
        """is_modified=False인 상세요구사항의 G열에는 MOD_FILL이 적용되지 않아야 한다."""
        details = [
            DetailRequirement(
                id="SFR-001-01",
                parent_id="SFR-001",
                category="기능요구사항",
                name="일반 명칭",
                content="일반 내용",
                order_index=0,
                is_modified=False,
            )
        ]
        state.set_detail(details)

        service = ExcelExportService()
        data = service.export(2)
        ws = _load_ws_from_bytes(data)

        cell = ws.cell(2, 7)
        # MOD_FILL이 아니어야 한다
        assert cell.fill.fgColor.rgb != MOD_FILL.fgColor.rgb, (
            "미수정 셀에 MOD_FILL이 적용되면 안 된다"
        )

    def test_modified_value_reflected_in_cell(self):
        """수정된 값이 G열에 실제로 반영되어야 한다."""
        details = [
            DetailRequirement(
                id="SFR-001-01",
                parent_id="SFR-001",
                category="기능요구사항",
                name="수정된 명칭",
                content="채팅으로 수정된 내용입니다.",
                order_index=0,
                is_modified=True,
            )
        ]
        state.set_detail(details)

        service = ExcelExportService()
        data = service.export(2)
        ws = _load_ws_from_bytes(data)

        assert ws.cell(2, 7).value == "채팅으로 수정된 내용입니다."


# ---------------------------------------------------------------------------
# UT-005-05: stage=2 + detailReqs 없음 → HTTPException 422
# ---------------------------------------------------------------------------


class TestStage2WithoutDetails:
    """UT-005-05: 상세요구사항 미생성 상태에서 stage=2 요청 시 422를 반환해야 한다."""

    def setup_method(self):
        state.reset_session()
        state.set_original(_make_originals())
        # 상세요구사항은 생성하지 않는다

    def test_export_stage2_without_details_raises_422(self):
        """상세요구사항이 없으면 HTTPException(422)이 발생해야 한다."""
        service = ExcelExportService()
        with pytest.raises(HTTPException) as exc_info:
            service.export(2)

        assert exc_info.value.status_code == 422
        detail = exc_info.value.detail
        assert detail["code"] == "DETAIL_NOT_GENERATED"

    def test_download_endpoint_stage2_without_details_returns_422(self):
        """GET /api/v1/download?stage=2에서 상세 미생성 시 422가 반환되어야 한다."""
        client = TestClient(app)
        response = client.get("/api/v1/download", params={"session_id": "test", "stage": 2})
        assert response.status_code == 422
